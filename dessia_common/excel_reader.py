#!/usr/bin/env python3
# -*- coding: utf-8 -*-
""" ExcelReader for DessiaObject. """

import inspect
from ast import literal_eval
from typing import Dict, List, Optional, Any, Tuple, TypeVar

import openpyxl

from dessia_common.utils.helpers import get_python_class_from_class_name


T = TypeVar("T")


def check_duplicate_attributes(attributes: List[str]) -> Optional[str]:
    """
    Check for the presence of duplicate attributes in a list of attribute names.

    If duplicates are found, return the first duplicated attribute name; otherwise, return None.
    """
    seen = set()
    for attribute in attributes:
        if attribute in seen:
            return attribute
        seen.add(attribute)
    return None


class ExcelDataExtract:
    """
    Data extracted from sheet of Excel file.

    :param class_module: module name extracted from cell R2C1.
    :param class_name: class name extracted from cell R2C2.
    :param attributes: List of attribute names extracted from row 3, starting from column 2.
    :param datas: Dictionary containing row-wise data (starting from row 4, column 2)
                  with column values as per attributes.
    """

    def __init__(self, class_module: str, class_name: str, attributes: List[str], datas: Dict[str, Any]):
        self.class_module = class_module
        self.class_name = class_name
        self.attributes = attributes
        self.datas = datas
        self.missed_attribute()

    @property
    def object_class(self) -> str:
        """ Get the full name of the class. """
        return f"{self.class_module}.{self.class_name}"

    @property
    def class_(self):
        """ Get the Python class object corresponding to the fully class name. """
        return get_python_class_from_class_name(self.object_class)

    @property
    def get_attributes_and_types(self) -> Dict[str, Any]:
        """
        Retrieve attributes and their corresponding types from the __init__ method of a class.

        Return a dictionary containing attributes and their respective types defined in the class constructor.
        """
        init_signature = inspect.signature(self.class_.__init__)
        init_parameters = init_signature.parameters

        attribute_info = {}

        for attr_name, param in init_parameters.items():
            if attr_name != "self":
                attr_info = {"type": param.annotation if param.annotation != param.empty else None,
                             "default_value": param.default if param.default != param.empty else "empty"}
                attribute_info[attr_name] = attr_info
        return attribute_info

    def missed_attribute(self):
        """ Checks for missing attributes in the current instance."""
        if len(self.attributes) != len(self.get_attributes_and_types.keys()):
            if len(self.attributes) > len(self.get_attributes_and_types.keys()):
                missing_attributes = set(self.get_attributes_and_types.keys()) - set(self.attributes)
            else:
                missing_attributes = set(self.get_attributes_and_types.keys()) - set(self.attributes)
            for attr_name in missing_attributes:
                if self.get_attributes_and_types[attr_name]["default_value"] == "empty":
                    raise ValueError(f"Missing attribute '{attr_name}' doesn't have a default value.")


class ExcelDatasExtracted:
    """
    Collection of data extracted from Excel.

    :param extracted_datas: A dictionary containing information extracted from the workbook sheets.
    Each key represents the sheet title, and its value the correspondinf ExcelDataExtract
    """

    def __init__(self, extracted_datas: Dict[str, ExcelDataExtract]):
        self.extracted_datas = extracted_datas


class ExcelReader:
    def __init__(self, stream):
        self.stream = stream
        self.workbook = openpyxl.load_workbook(stream, data_only=True)
        self.sheet_titles = [sheet.title for sheet in self.workbook.worksheets]
        self.main_sheet = self.workbook.worksheets[0].title

    def get_location(self, cell: openpyxl.cell.cell.Cell) -> Tuple[str, str]:
        """
        Retrieve the location of the given cell's hyperlink.

        Return tuple containing the sheet name and cell address referred to by the hyperlink.
        """
        if cell.hyperlink.location:
            target = cell.hyperlink.location.split('!')
        elif cell.hyperlink.target:
            target = cell.hyperlink.target.split('!')

        else:
            raise ValueError("The provided cell does not contain a valid hyperlink location or target.")

        sheet_name = target[0].replace("#", '')
        if sheet_name not in self.sheet_titles:
            raise ValueError(f"The sheet '{sheet_name}' referenced by the hyperlink does not exist.")
        return sheet_name, target[1]

    def update_attribute_values(self, cell, values):
        """
        Updates attribute values based on the provided cell and values.

        - Depending on the data type of the cell it returns
          - For 'Dict' type: a dictionary with keys extracted from the cell
            and replaced values from the 'values' parameter.
          - For other types (excluding 'List', 'Set', 'Tuple'): Replaces the values with
            the first element of 'values' if 'values' is a list and 'cell.value' does not
            contain any of the container types.
          - Otherwise, 'values' itself.
        """
        container = ["Dict", "List", "Set", "Tuple"]
        target = self.get_location(cell)
        dict_keys = []
        if "Dict" in cell.value:
            actual_value = {}
            sheet = self.workbook[target]
            column = sheet["A"]
            for row in column[3:]:
                dict_keys.append(row.value.split('.')[1])

            for key, value in zip(dict_keys, values):
                actual_value[key] = value

            return actual_value

        if isinstance(values, list) and (not any(type_ in cell.value for type_ in container)):
            actual_value = values[0]
        else:
            actual_value = values
        return actual_value

    @staticmethod
    def get_data(values: List[T], attributes: List[str], init_attributes: List[str]) -> Dict[str, T]:
        """
        Extracts relevant data based on attributes and initial attributes for object instantiation.

        :param values: List containing values corresponding to attributes.
        :param attributes: List of attributes related to the object.
        :param init_attributes: List of initial attributes for object instantiation.

        :return: Dictionary containing extracted object data based on matching attributes and initial attributes.
        """
        object_data = {}
        for attribute, value in zip(attributes, values):
            if attribute in init_attributes:
                if isinstance(value, str):
                    try:
                        object_data[attribute] = literal_eval(value)
                    except (ValueError, SyntaxError):
                        object_data[attribute] = value
                else:
                    object_data[attribute] = value
        return object_data

    def create_object(self, extracted_data, value_set):
        """ Create an object of a specified class using provided data. """
        object_data = self.get_data(value_set, extracted_data.attributes, extracted_data.get_attributes_and_types)
        object_ = extracted_data.class_(**object_data)
        if not object_.name:
            object_.name = ""

        return object_

    def instantiate_main_object(self, instantiated_objects, extracted_data: ExcelDataExtract, key: str):
        """
        Instantiate main objects based on provided values and attributes.

        :param instantiated_objects: Dictionary containing instantiated objects.
        :param key: Key to identify the instantiated objects.
        :param extracted_data: Object containing module name, class name, and attributes.

        :return: Updated dictionary of instantiated objects.
        """
        objects = []

        value_set = list(extracted_data.datas.values())[0]
        for index, value in enumerate(value_set):
            if isinstance(value, openpyxl.cell.cell.Cell):
                replaced_value = instantiated_objects[self.get_location(value)[0]]
                value_set[index] = self.update_attribute_values(value, replaced_value)

        objects.append(self.create_object(extracted_data, value_set))
        instantiated_objects[key] = objects
        return instantiated_objects

    def process_sub_objects(self, cell: openpyxl.cell.cell.Cell, workbook: openpyxl.Workbook):
        """
        Process sub-objects based on cell data.

        :param cell: The cell containing data.
        :param workbook: The workbook containing the cell.

        :return: object or any. The instantiated sub-object or original cell value.
        """
        if isinstance(cell, openpyxl.cell.cell.Cell):
            sheet_title, row_target = self.get_location(cell)
            sub_module_name = workbook[sheet_title]["A2"].value
            sub_class_name = workbook[sheet_title]["B2"].value
            sheet_target = workbook[sheet_title]
            sub_attributes = list(sheet_target.iter_rows(min_row=3, max_row=3, min_col=2, values_only=True))[0]

            moment_values = [c.value for c in sheet_target[int(row_target[1:])][1:]]
            extracted_data = ExcelDataExtract(class_module=sub_module_name, class_name=sub_class_name,
                                              attributes=sub_attributes, datas=None)

            return self.create_object(extracted_data, moment_values)
        return cell

    def process_multiple_hyperlink_rows(self, instantiated_objects: Dict, extracted_data: ExcelDataExtract, key: str):
        """
        Process multiple rows in a sheet, each representing an object with hyperlinks.

        :param instantiated_objects: Dictionary containing instantiated objects.
        :param key: Key to associate with the list of instantiated objects.
        :param extracted_data: Object containing information to process.

        :return: Updated dictionary of instantiated objects.
        """

        objects = []
        for value_set in extracted_data.datas.values():
            sub_values = [self.process_sub_objects(cell, self.workbook) for cell in value_set]
            objects.append(self.create_object(extracted_data, sub_values))

        instantiated_objects[key] = objects
        return instantiated_objects

    def process_single_hyperlink_row(self, instantiated_objects: Dict, extracted_data: ExcelDataExtract, key: str):
        """
        Process a single row in a sheet representing an object with hyperlinks.

        :param instantiated_objects: Dictionary containing instantiated objects.
        :param key: Key to associate with the list of instantiated objects.
        :param extracted_data: Object containing information to process.

        :return: Updated dictionary of instantiated objects.
        """
        objects = []
        value_set = list(extracted_data.datas.values())[0]
        for k, cell in enumerate(value_set):
            if isinstance(cell, openpyxl.cell.cell.Cell):
                val_replace = instantiated_objects[self.get_location(cell)[0]]
                value_set[k] = self.update_attribute_values(cell, val_replace)

        objects.append(self.create_object(extracted_data, value_set))

        instantiated_objects[key] = objects
        return instantiated_objects

    def process_simple_sheet(self, instantiated_objects: Dict, extracted_data: ExcelDataExtract, key: str):
        """
        Process a sheet without hyperlinks containing simple cell values to instantiate objects.

        :param instantiated_objects: A dictionary containing instantiated objects.
        :param key: The key representing the current sheet being processed.
        :param extracted_data: A list of values containing module name, class name, attributes, and data.

        :return: The updated dictionary of instantiated objects after processing the sheet.
        """
        objects = []
        for value_set in extracted_data.datas.values():
            objects.append(self.create_object(extracted_data, value_set))

        instantiated_objects[key] = objects
        return instantiated_objects

    def process_workbook(self) -> ExcelDatasExtracted:
        """
        Process the workbook's sheets and extract information.

        This function iterates through each sheet in the workbook and gathers specific data from each sheet.
        It collects module and class names along with attributes and associated data.
        """
        extracted_datas = {}
        for sheet in self.workbook.worksheets:
            class_info = (sheet.cell(row=2, column=1).value, sheet.cell(row=2, column=2).value)

            attributes = []
            for value in sheet.iter_cols(min_row=3, max_row=3, min_col=2, values_only=True):
                attributes.append(value[0])

            row_data = {}
            for i, value in enumerate(sheet.iter_rows(min_row=4, min_col=2, values_only=True)):
                column_values = []
                for j, _ in enumerate(value):
                    cell_value = sheet.cell(row=i + 4, column=j + 2)
                    if cell_value.hyperlink:
                        column_values.append(cell_value)
                    else:
                        column_values.append(cell_value.value)
                row_data[i] = column_values

            extracted_datas[sheet.title] = ExcelDataExtract(class_module=class_info[0], class_name=class_info[1],
                                                            attributes=attributes,
                                                            datas=row_data.copy())
        return ExcelDatasExtracted(extracted_datas=extracted_datas)

    def read_workbook(self):
        """
        Read and process data to instantiate objects from the workbook.

        Reads the processed workbook data, instantiates objects, and organizes them based on the provided classes and
        modules.

        Note:
        This method does not support processing variables of complex types, such as `List[Tuple[float, float]]`, due to
        limitations in interpreting complex nested structures within Excel sheets.

        :return: The instantiated object obtained from the main sheet's data.
        """
        instantiated_objects = {}
        data_structure = self.process_workbook()
        stack = list(data_structure.extracted_datas.items())

        while stack:
            key, extracted_data = stack.pop()
            if key == self.main_sheet:
                instantiated_objects = self.instantiate_main_object(instantiated_objects, extracted_data, key)
                break

            is_cell_instance = any((any(isinstance(cell_value, openpyxl.cell.cell.Cell) for cell_value in val) for val
                                    in extracted_data.datas.values()))
            if is_cell_instance:
                nb_objects_in_sheet = len(extracted_data.datas.keys())
                if nb_objects_in_sheet > 1:
                    instantiated_objects = self.process_multiple_hyperlink_rows(instantiated_objects,
                                                                                extracted_data, key)

                else:
                    instantiated_objects = self.process_single_hyperlink_row(instantiated_objects, extracted_data, key)
            else:
                instantiated_objects = self.process_simple_sheet(instantiated_objects, extracted_data, key)
        return instantiated_objects[self.main_sheet][0]

    def process_workbook_catalog(self) -> ExcelDatasExtracted:
        """
        Process the workbook's sheets and extract information (Catalog).

        This function iterates through each sheet in the workbook and gathers specific data from each sheet.
        It collects module and class names along with attributes and associated data.
        """

        extracted_datas = {}
        for sheet in self.workbook.worksheets:
            class_info = (sheet.cell(row=1, column=1).value, sheet.cell(row=1, column=2).value)

            attributes = []
            for value in sheet.iter_cols(min_row=2, max_row=2, min_col=1, values_only=True):
                attributes.append(value[0])

            duplicate_attribute = check_duplicate_attributes(attributes)
            if duplicate_attribute:
                raise ValueError(f"Duplicate attribute '{duplicate_attribute}' detected. Each attribute name should "
                                 f"be unique.")

            row_data = {}
            for i, value in enumerate(sheet.iter_rows(min_row=3, min_col=1, values_only=True)):
                column_values = []
                for j, _ in enumerate(value):
                    cell_value = sheet.cell(row=i + 3, column=j + 1)
                    column_values.append(cell_value.value)
                row_data[i] = column_values

            extracted_datas[sheet.title] = ExcelDataExtract(class_module=class_info[0], class_name=class_info[1],
                                                            attributes=attributes,
                                                            datas=row_data.copy())
        return ExcelDatasExtracted(extracted_datas=extracted_datas)

    def read_catalog(self):
        """
        Read and process data to instantiate objects from the workbook (Catalog).

        This method retrieves cell values from a workbook, processes them, and creates instantiated
        objects based on the processed data.

        :return: A dictionary containing instantiated objects created from the processed workbook data.
        """
        instantiated_objects = {}
        data_structure = self.process_workbook_catalog()
        stack = list(data_structure.extracted_datas.items())

        while stack:
            key, extracted_data = stack.pop()
            instantiated_objects = self.process_simple_sheet(instantiated_objects, extracted_data, key)
        return instantiated_objects

    def close(self):
        """
        Closes the workbook.

        This method closes the associated workbook, ensuring any changes made are saved and resources are released.
        """
        self.workbook.close()
