#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

"""
import builtins
import sys
import warnings
import math
import random
import copy
from functools import reduce
import collections
from copy import deepcopy
import inspect
import json
import bson
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, PatternFill
from openpyxl import Workbook
import openpyxl.utils

from typing import List, Dict, Type, Tuple, Union, Any, \
    get_type_hints, get_origin, get_args
try:
    from typing import TypedDict  # >=3.8
except ImportError:
    from mypy_extensions import TypedDict  # <=3.7
import traceback as tb
from dessia_common.typings import Measure, JsonSerializable,\
    Subclass, InstanceOf

from importlib import import_module


JSONSCHEMA_HEADER = {"definitions": {},
                     "$schema": "http://json-schema.org/draft-07/schema#",
                     "type": "object",
                     "required": [],
                     "properties": {}}

TYPING_EQUIVALENCES = {int: 'number', float: 'number',
                       bool: 'boolean', str: 'string'}

TYPES_STRINGS = {int: 'int', float: 'float', bool: 'boolean', str: 'str',
                 list: 'list', tuple: 'tuple', dict: 'dict'}

SEQUENCE_TYPINGS = ['List', 'Sequence', 'Iterable']

_FORBIDDEN_ARGNAMES = ['self', 'cls', 'progress_callback', 'return']

TYPES_FROM_STRING = {'unicode': str, 'str': str, 'float': float,
                     'int': int, 'bool': bool}


class ExceptionWithTraceback(Exception):
    def __init__(self, message, traceback_=''):
        self.message = message
        self.traceback = traceback_

    def __str__(self):
        return '{}\nTraceback:\n{}'.format(self.message, self.traceback)


class DeepAttributeError(ExceptionWithTraceback, AttributeError):
    pass


class ModelError(Exception):
    pass


class ConsistencyError(Exception):
    pass


class SerializationError(Exception):
    pass


class DeserializationError(Exception):
    pass


class UntypedArgumentError(Exception):
    pass


# DEPRECATED_ATTRIBUTES = {'_editable_variss' : '_allowed_methods'}
def deprecated(use_instead=None):
    def decorated(function):
        def wrapper(*args, **kwargs):
            deprecation_warning(function.__name__, 'Function', use_instead)
            print('Traceback : ')
            tb.print_stack(limit=2)
            return function(*args, **kwargs)

        return wrapper

    return decorated


def deprecation_warning(name, object_type, use_instead=None):
    warnings.simplefilter('once', DeprecationWarning)
    msg = "\n\n{} {} is deprecated.\n".format(object_type, name)
    msg += "It will be removed in a future version.\n"
    if use_instead is not None:
        msg += "Use {} instead.\n".format(use_instead)
    warnings.warn(msg, DeprecationWarning)
    return msg


class DessiaObject:
    """
    Base class for Dessia's platform compatible objects.
    Gathers generic methods and attributes

    :cvar bool _standalone_in_db:
        Indicates wether class objects should be independant
        in database or not.
        If False, object will only exist inside its parent.
    :cvar bool _eq_is_data_eq:
        Indicates which type of equality check is used:
        strict equality or equality based on data. If False, Python's
        object __eq__ method is used (ie. strict),
        else, user custom data_eq is used (ie. data)
    :cvar List[str] _non_serializable_attributes:
        [Advanced] List of instance attributes that should not
        be part of serialization with to_dict method. These will not
        be displayed in platform object tree, for instance.
    :cvar List[str] _non_data_eq_attributes:
        [Advanced] List of instance attributes that should not
        be part of equality check with data__eq__ method
        (if _eq_is_data_eq is True).
    :cvar List[str] _non_data_hash_attributes:
        [Advanced] List of instance attributes that should not
        be part of hash computation with data__hash__ method
        (if _eq_is_data_eq is True).
    :cvar List[str] _ordered_attributes:
        Documentation not available yet.
    :cvar List[str] _titled_attributes:
        Documentation not available yet.
    :cvar List[str] _init_variables:
        Documentation not available yet.
    :cvar List[str] _export_formats:
        List of all available export formats. Class must define a
        export_[format] for each format in _export_formats
    :cvar List[str] _allowed_methods:
        List of all methods that are runnable from platform.
    :cvar List[str] _whitelist_attributes:
        Documentation not available yet.
    :cvar List[str] _whitelist_attributes: List[str]


    :ivar str name: Name of object.
    :ivar Any **kwargs: Additionnal user metadata
    """
    _standalone_in_db = False
    _non_serializable_attributes = []
    _non_editable_attributes = []
    _non_data_eq_attributes = ['name']
    _non_data_hash_attributes = ['name']
    _ordered_attributes = []
    _titled_attributes = []
    _eq_is_data_eq = True
    
    _init_variables = None
    _allowed_methods = []
    _whitelist_attributes = []

    def __init__(self, name: str = '', **kwargs):
        self.name = name
        for property_name, property_value in kwargs.items():
            setattr(self, property_name, property_value)

    def __hash__(self):
        if self._eq_is_data_eq:
            return self._data_hash()
        else:
            return object.__hash__(self)

    def __eq__(self, other_object):
        if self._eq_is_data_eq:
            if self.__class__.__name__ != other_object.__class__.__name__:
                return False
            if self._data_hash() != other_object._data_hash():
                return False
            return self._data_eq(other_object)
        return object.__eq__(self, other_object)

    def _data_eq(self, other_object):
        if full_classname(self) != full_classname(other_object):
            return False

        eq_dict = self._serializable_dict()
        if 'name' in eq_dict:
            del eq_dict['name']
            
        other_eq_dict = other_object._serializable_dict()

        for key, value in eq_dict.items():
            other_value = other_eq_dict[key]
            if value != other_value:
                return False
        return True

    def _data_hash(self):
        hash_ = 0
        forbidden_keys = (self._non_data_eq_attributes
                          + self._non_data_hash_attributes
                          + ['package_version', 'name'])
        for key, value in self._serializable_dict().items():
            if key not in forbidden_keys:
                if isinstance(value, list):
                    hash_ += list_hash(value)
                elif isinstance(value, dict):
                    hash_ += dict_hash(value)
                elif isinstance(value, str):
                    hash_ += sum([ord(v) for v in value])
                else:
                    hash_ += hash(value)
        return int(hash_ % 1e5)

    def _data_diff(self, other_object):
        """
        Make a diff between two objects
        returns: different values, missing keys in other object
        """
        missing_keys_in_other_object = []
        diff_values = {}
        
        # eq_dict = {k: v for k, v in self.to_dict().items()
        #            if (k not in ['package_version', 'name'])\
        #                and (k not in self._non_data_eq_attributes)}
        eq_dict = self._serializable_dict()
        # other_eq_dict = other_object.to_dict()
        other_eq_dict = other_object._serializable_dict()

        for key, value in eq_dict.items():
            if key not in other_eq_dict:
                missing_keys_in_other_object.append(key)
            else:                
                other_value = other_eq_dict[key]
                if value != other_value:
                    diff_values[key] = (value, other_value)
                
        return diff_values, missing_keys_in_other_object

    @property
    def full_classname(self):
        return full_classname(self)

    def base_dict(self):
        package_name = self.__module__.split('.')[0]
        if package_name in sys.modules:
            package = sys.modules[package_name]
            if hasattr(package, '__version__'):
                package_version = package.__version__
            else:
                package_version = None
        else:
            package_version = None

        object_class = self.__module__ + '.' + self.__class__.__name__
        dict_ = {'name': self.name, 'object_class': object_class}
        if package_version:
            dict_['package_version'] = package_version
        return dict_

    def _serializable_dict(self):

        dict_ = {k: v for k, v in self.__dict__.items()
                 if k not in self._non_serializable_attributes
                 and not k.startswith('_')}
        return dict_


    def to_dict(self) -> JsonSerializable:
        """
        Generic to_dict method
        """
        dict_ = self._serializable_dict()
        if hasattr(self, 'Dict'):
            # !!! This prevent us to call DessiaObject.to_dict()
            # from an inheriting object which implement a Dict method,
            # because of the infinite recursion it creates.
            deprecation_warning(name='Dict', object_type='Function',
                                use_instead='to_dict')
            serialized_dict = self.Dict()
        else:
            # Default to dict
            serialized_dict = self.base_dict()
            serialized_dict.update(serialize_dict(dict_))
        # serialized_dict['hash'] = self.__hash__()
        return serialized_dict

    @classmethod
    def dict_to_object(cls, dict_: JsonSerializable,
                       force_generic: bool = False) -> 'DessiaObject':
        """
        Generic dict_to_object method
        """
        if hasattr(cls, 'DictToObject'):
            deprecation_warning(name='DictToObject', object_type='Function',
                                use_instead='dict_to_object')
            return cls.DictToObject(dict_)

        if cls is not DessiaObject:
            obj = dict_to_object(dict_=dict_, class_=cls,
                                 force_generic=force_generic)
            return obj
        elif 'object_class' in dict_:
            obj = dict_to_object(dict_=dict_, force_generic=force_generic)
            return obj
        else:
            # Using default
            # TODO: use jsonschema
            raise NotImplementedError('No object_class in dict')

    @classmethod
    def base_jsonschema(cls):
        jsonschema = deepcopy(JSONSCHEMA_HEADER)
        jsonschema['properties']['name'] = {
            'type': 'string',
            "title": "Object Name",
            "description": "Object name",
            "editable": True,
            "default_value": "Object Name"
        }
        return jsonschema

    @classmethod
    def jsonschema(cls):
        if hasattr(cls, '_jsonschema'):
            _jsonschema = cls._jsonschema
            return _jsonschema

        # Get __init__ method and its annotations
        init = cls.__init__
        if cls._init_variables is None:
            annotations = get_type_hints(init)
        else:
            annotations = cls._init_variables

        # Get ordered variables
        if cls._ordered_attributes:
            ordered_attributes = cls._ordered_attributes
        else:
            ordered_attributes = list(annotations.keys())

        unordered_count = 0

        # Parse docstring
        try:
            parsed_docstring = parse_docstring(cls)
        except Exception:
            parsed_docstring = {
                'description': 'Docstring parsing failed',
                'attributes': {}
            }
        parsed_attributes = parsed_docstring['attributes']

        # Initialize jsonschema
        _jsonschema = deepcopy(JSONSCHEMA_HEADER)

        required_arguments, default_arguments = inspect_arguments(method=init,
                                                                  merge=False)
        _jsonschema['required'] = required_arguments
        _jsonschema['standalone_in_db'] = cls._standalone_in_db
        _jsonschema['description'] = parsed_docstring['description']
        _jsonschema['python_typing'] = str(cls)

        # Set jsonschema
        for annotation in annotations.items():
            name = annotation[0]
            if name in ordered_attributes:
                order = ordered_attributes.index(name)
            else:
                order = len(ordered_attributes) + unordered_count
                unordered_count += 1
            if name in cls._titled_attributes:
                title = cls._titled_attributes[name]
            else:
                title = None

            if name != 'return':
                editable = name not in cls._non_editable_attributes
                annotation_type = type_from_annotation(annotation[1], cls)
                annotation = (annotation[0], annotation_type)
                jss_elt = jsonschema_from_annotation(
                    annotation=annotation, jsonschema_element={},
                    order=order, editable=editable, title=title
                )
                if name in parsed_attributes:
                    description = parsed_attributes[name]['desc']
                    jss_elt[name]['description'] = description
                _jsonschema['properties'].update(jss_elt)
                if name in default_arguments.keys():
                    default = set_default_value(_jsonschema['properties'],
                                                name,
                                                default_arguments[name])
                    _jsonschema['properties'].update(default)

        _jsonschema['classes'] = [cls.__module__ + '.' + cls.__name__]
        _jsonschema['whitelist_attributes'] = cls._whitelist_attributes
        return _jsonschema

    @property
    def _method_jsonschemas(self):
        """
        Generates dynamic jsonschemas for methods of class
        """
        jsonschemas = {}
        class_ = self.__class__

        # TOCHECK Backward compatibility. Will need to be changed
        if hasattr(class_, '_dessia_methods'):
            allowed_methods = class_._dessia_methods
        else:
            allowed_methods = class_._allowed_methods

        valid_method_names = [m for m in dir(class_)
                              if not m.startswith('_')
                              and m in allowed_methods]

        for method_name in valid_method_names:
            method = getattr(class_, method_name)

            if not isinstance(method, property):
                required_args, default_args = inspect_arguments(method=method,
                                                                merge=False)
                annotations = get_type_hints(method)
                if annotations:
                    jsonschemas[method_name] = deepcopy(JSONSCHEMA_HEADER)
                    jsonschemas[method_name]['required'] = []
                    jsonschemas[method_name]['method'] = True
                    for i, annotation in enumerate(annotations.items()):
                        # TOCHECK Not actually ordered
                        argname = annotation[0]
                        if argname not in _FORBIDDEN_ARGNAMES:
                            if argname in required_args:
                                jsonschemas[method_name]['required'].append(
                                    str(i))
                            jsonschema_element = \
                                jsonschema_from_annotation(annotation, {}, i)[
                                    argname]

                            jsonschemas[method_name]['properties'][
                                str(i)] = jsonschema_element
                            if argname in default_args.keys():
                                default = set_default_value(
                                    jsonschemas[method_name]['properties'],
                                    str(i),
                                    default_args[argname])
                                jsonschemas[method_name]['properties'].update(
                                    default)
        return jsonschemas

    def method_dict(self, method_name=None, method_jsonschema=None):
        if method_name is None and method_jsonschema is None:
            msg = 'No method name not jsonschema provided'
            raise NotImplementedError(msg)

        if method_name is not None and method_jsonschema is None:
            method_jsonschema = self._method_jsonschemas[method_name]

        dict_ = default_dict(method_jsonschema)
        return dict_

    def dict_to_arguments(self, dict_, method):
        method_object = getattr(self, method)
        args_specs = inspect.getfullargspec(method_object)
        allowed_args = args_specs.args[1:]

        arguments = {}
        for i, arg in enumerate(allowed_args):
            if str(i) in dict_:
                arg_specs = args_specs.annotations[arg]
                value = dict_[str(i)]
                try:
                    deserialized_value = deserialize_argument(arg_specs, value)
                except TypeError:
                    msg = 'Error in deserialisation of value: '
                    msg += '{} of expected type {}'.format(value, arg_specs)
                    raise TypeError(msg)
                arguments[arg] = deserialized_value
        return arguments

    def save_to_file(self, filepath, indent=0):
        if isinstance(filepath, str):
            if not filepath.endswith('.json'):
                filepath += '.json'
            file = open(filepath, 'w')
        else:
            file = filepath
        json.dump(self.to_dict(), file, indent=indent)
        
        if isinstance(filepath, str):
            file.close()    
        
    @classmethod
    def load_from_file(cls, filepath):
        if isinstance(filepath, str):
            with open(filepath, 'r') as file:
                dict_ = json.load(file)
        else:
            dict_ = json.loads(filepath.read().decode('utf-8'))
        return cls.dict_to_object(dict_)

    def is_valid(self):
        return True

    def copy(self, deep=True):
        if deep:
            return self.__deepcopy__()
        else:
            return self.__copy__()

    def __copy__(self):
        """
        Generic copy use inits of objects
        """
        class_argspec = inspect.getfullargspec(self.__class__)
        dict_ = {}
        for arg in class_argspec.args:
            if arg != 'self':
                value = self.__dict__[arg]

                if hasattr(value, '__copy__'):
                    dict_[arg] = value.__copy__()
                else:
                    dict_[arg] = value
        return self.__class__(**dict_)

    def __deepcopy__(self, memo=None):
        """
        Generic deep copy use inits of objects
        """
        class_argspec = inspect.getfullargspec(self.__class__)
        if memo is None:
            memo = {}
        dict_ = {}
        for arg in class_argspec.args:
            if arg != 'self':
                dict_[arg] = deepcopy_value(getattr(self, arg), memo=memo)
        return self.__class__(**dict_)

    def volmdlr_volume_model(self, **kwargs):
        if hasattr(self, 'volmdlr_primitives'):
            import volmdlr as vm  # !!! Avoid circular imports, is this OK ?
            if hasattr(self, 'volmdlr_primitives_step_frames'):
                return vm.core.MovingVolumeModel(self.volmdlr_primitives(**kwargs),
                                            self.volmdlr_primitives_step_frames(**kwargs))
            # else:
            #     if frame is None:
            #         frame = vm.OXYZ
            # try:
            return vm.core.VolumeModel(self.volmdlr_primitives(**kwargs))
            # except TypeError:
            #     return vm.core.VolumeModel(self.volmdlr_primitives())
        msg = 'Object of type {} does not implement volmdlr_primitives'
        raise NotImplementedError(msg.format(self.__class__.__name__))


    def plot(self, **kwargs):
        """

        """
        if hasattr(self, 'plot_data'):
            import plot_data
            for data in self.plot_data(**kwargs):
                plot_data.plot_canvas(plot_data_object=data,
                                      canvas_id='canvas',
                                      width=1400, height=900,
                                      debug_mode=False)
        else:
            msg = 'Class {} does not implement a plot_data method' \
                  'to define what to plot'
            raise NotImplementedError(msg.format(self.__class__.__name__))

    def mpl_plot(self, **kwargs):
        axs = []
        if hasattr(self, 'plot_data'):
            try:
                plot_datas = self.plot_data(**kwargs)
            except TypeError as error:
                raise TypeError('{}.{}'.format(self.__class__.__name__, error))
            for data in plot_datas:
                if hasattr(data, 'mpl_plot'):
                    ax = data.mpl_plot()
                    axs.append(ax)
        else:
            msg = 'Class {} does not implement a plot_data method' \
                  'to define what to plot'
            raise NotImplementedError(msg.format(self.__class__.__name__))

        return axs

    def babylonjs(self, use_cdn=True, debug=False, **kwargs):
        self.volmdlr_volume_model(**kwargs).babylonjs(use_cdn=use_cdn, debug=debug)

    def _displays(self, **kwargs) -> List[JsonSerializable]:
        if hasattr(self, '_display_angular'):
            # Retro-compatibility
            deprecation_warning(name='_display_angular', object_type='method',
                                use_instead='display_angular')
            return self._display_angular(**kwargs)

        if 'reference_path' in kwargs:
            reference_path = kwargs['reference_path']
        else:
            reference_path = ''
        displays = []
        if hasattr(self, 'babylon_data'):
            display_ = DisplayObject(type_='cad', data=self.babylon_data(),
                                     reference_path=reference_path)
            displays.append(display_.to_dict())
        elif hasattr(self, 'volmdlr_primitives')\
                or (self.__class__.volmdlr_volume_model
                    is not DessiaObject.volmdlr_volume_model):
            model = self.volmdlr_volume_model()
            display_ = DisplayObject(type_='cad', data=model.babylon_data(),
                                     reference_path=reference_path)
            displays.append(display_.to_dict())
        if hasattr(self, 'plot_data'):
            plot_data = self.plot_data()
            if is_sequence(plot_data):
                for plot in plot_data:
                    display_ = DisplayObject(type_='plot_data', data=plot,
                                             reference_path=reference_path)
                    displays.append(display_.to_dict())
            else:
                msg = 'plot_data must return a sequence. Found {}'
                raise ValueError(msg.format(type(plot_data)))
        if hasattr(self, 'to_markdown'):
            display_ = DisplayObject(type_='markdown', data=self.to_markdown(),
                                     reference_path=reference_path)
            displays.append(display_.to_dict())
        return displays

    def _check_platform(self):
        """
        Reproduce lifecycle on platform (serialization, display)
        """
        self.dict_to_object(json.loads(json.dumps(self.to_dict())))
        bson.BSON.encode(self.to_dict())
        json.dumps(self._displays())

    def to_xlsx(self, filepath):
        max_column_width = 40
        color_dessIA1 = "95B3D8"
        color_dessIA2 = "78909C"
        grey1 = "CCCCCC"
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        pattern_color1 = PatternFill(
            fill_type="solid",
            start_color=color_dessIA1,
            end_color=color_dessIA1)

        pattern_color2 = PatternFill(
            fill_type="solid",
            start_color=color_dessIA2,
            end_color=color_dessIA2)



        wb = Workbook()
        # grab the active worksheet
        ws1 = wb.active
        ws1.title = 'Object {}'.format(self.__class__.__name__)

        ws1['A1'] = 'Module'
        ws1['B1'] = 'Class'
        ws1['C1'] = 'name'

        ws1['A1'].border = thin_border
        ws1['B1'].border = thin_border
        ws1['C1'].border = thin_border
        ws1['A1'].fill = pattern_color1
        ws1['B1'].fill = pattern_color1
        ws1['C1'].fill = pattern_color1


        ws1['A2'] = self.__module__
        ws1['B2'] = self.__class__.__name__
        ws1['C2'] = self.name

        ws1['A2'].border = thin_border
        ws1['B2'].border = thin_border
        ws1['C2'].border = thin_border
        ws1['A2'].fill = pattern_color1
        ws1['B2'].fill = pattern_color1
        ws1['C2'].fill = pattern_color1



        ws1['A4'] = 'Attribute'
        ws1['A5'] = 'Value'
        ws1['A4'].border = thin_border
        ws1['A5'].border = thin_border


        # name_column_width = 0
        i = 1
        for (k, v) in sorted(self.__dict__.items()):
            if (not k.startswith('_')) and k != 'name':
                cell1 = ws1.cell(row=4, column=i, value=k)
                cell2 = ws1.cell(row=5, column=i, value=str(v))

                cell1.border = thin_border
                cell1.fill = pattern_color2
                cell2.border = thin_border

                i += 1

                column_width = min((len(k) + 1.5), max_column_width)
                column_name = openpyxl.utils.cell.get_column_letter(i)
                ws1.column_dimensions[column_name].width = column_width

        wb.save("{}.xlsx".format(filepath))

    def to_step(self, filepath):
        """
        filepath can be a str or an io.StringIO
        """
        return self.volmdlr_volume_model().to_step(filepath=filepath)

    def _export_formats(self):
        formats = [('json', 'save_to_file'),
                   ('xlsx', 'to_xlsx')]
        if hasattr(self, 'volmdlr_primitives'):
            formats.append(('step', 'to_step'))
        return formats


class Catalog(DessiaObject):
    def __init__(self, objects: List[DessiaObject], name: str = ''):
        self.objects = objects
        DessiaObject.__init__(self, name=name)


class DisplayObject(DessiaObject):
    def __init__(self, type_: str,
                 data: Union[JsonSerializable, DessiaObject],
                 reference_path: str = '', name: str = ''):
        self.type_ = type_
        self.data = data

        self.reference_path = reference_path
        DessiaObject.__init__(self, name=name)


class Parameter(DessiaObject):
    def __init__(self, lower_bound, upper_bound, periodicity=None, name=''):
        DessiaObject.__init__(self, name=name,
                              lower_bound=lower_bound,
                              upper_bound=upper_bound,
                              periodicity=periodicity)

    def random_value(self):
        return random.uniform(self.lower_bound, self.upper_bound)

    def are_values_equal(self, value1, value2, tol=1e-2):
        if self.periodicity is not None:
            value1 = value1 % self.periodicity
            value2 = value2 % self.periodicity

        return math.isclose(value1, value2, abs_tol=tol)

    def normalize(self, value):
        normalized_value = (value - self.lower_bound) / (self.upper_bound - self.lower_bound)
        return normalized_value

    def original_value(self, normalized_value):
        value = normalized_value * (self.upper_bound - self.lower_bound) + self.lower_bound
        return value

    def optimizer_bounds(self):
        if self.periodicity is not None:
            return (self.lower_bound - 0.5 * self.periodicity,
                    self.upper_bound + 0.5 * self.periodicity)


class ParameterSet(DessiaObject):
    def __init__(self, values, name=''):
        self.values = values

        DessiaObject.__init__(self, name=name)

    @property
    def parameters(self):
        parameters = [Parameter(min(v), max(v), name=k)
                      for k, v in self.values.items()]
        return parameters

    @property
    def means(self):
        means = {k: sum(v)/len(v) for k, v in self.values.items()}
        return means


class Filter(TypedDict):
    attribute: str
    operator: str
    bound: float


class Evolution(DessiaObject):
    """
    Defines a generic evolution

    :param evolution: float list
    :type evolution: list
    """
    _non_data_eq_attributes = ['name']
    _non_data_hash_attributes = ['name']
    _generic_eq = True

    def __init__(self, evolution: List[float] = None, name: str = ''):
        if evolution is None:
            evolution = []
        self.evolution = evolution

        DessiaObject.__init__(self, name=name)

    def _displays(self):
        displays = [{'angular_component': 'app-evolution1d',
                     'table_show': False,
                     'evolution': [self.evolution],
                     'label_y': ['evolution']}]
        return displays

    def update(self, evolution):
        """
        Update the evolution list
        """
        self.evolution = evolution


class CombinationEvolution(DessiaObject):
    _non_data_eq_attributes = ['name']
    _non_data_hash_attributes = ['name']
    _generic_eq = True

    def __init__(self, evolution1: List[Evolution],
                 evolution2: List[Evolution], title1: str = 'x',
                 title2: str = 'y', name: str = ''):

        self.evolution1 = evolution1
        self.evolution2 = evolution2

        self.x_, self.y_ = self.genere_xy()

        self.title1 = title1
        self.title2 = title2

        DessiaObject.__init__(self, name=name)

    def _displays(self):
        displays = [{
            'angular_component': 'app-evolution2d-combination-evolution',
            'table_show': False,
            'evolution_x': [self.x_], 'label_x': ['title1'],
            'evolution_y': [self.y_], 'label_y': ['title2']
        }]
        return displays

    def update(self, evol1, evol2):
        """
        Update the CombinationEvolution object

        :param evol1: list
        :param evol2: list
        """
        for evolution, ev1 in zip(self.evolution1, evol1):
            evolution.update(ev1)
        for evolution, ev2 in zip(self.evolution2, evol2):
            evolution.update(ev2)
        self.x_, self.y_ = self.genere_xy()

    def genere_xy(self):
        x, y = [], []
        for evol in self.evolution1:
            x = x + evol.evolution
        for evol in self.evolution2:
            y = y + evol.evolution
        return x, y


def number2factor(number):
    """
    Temporary function : Add to some tools package
    Finds all the ways to combine elements
    """
    factor_range = range(1, int(number ** 0.5) + 1)

    if number:
        factors = list(set(reduce(list.__add__, ([i, number // i]
                                                 for i in factor_range
                                                 if number % i == 0))))

        grids = [(factor_x, int(number / factor_x))
                 for factor_x in factors
                 if (number / factor_x).is_integer()]
    else:
        grids = []
    return grids


def number3factor(number, complete=True):
    """
    Temporary function : Add to some tools package
    Finds all the ways to combine elements
    """
    factor_range = range(1, int(number ** 0.5) + 1)

    if number:
        factors = list(set(reduce(list.__add__, ([i, number // i]
                                                 for i in factor_range
                                                 if number % i == 0))))
        if not complete:
            grids = get_incomplete_factors(number, factors)

        else:
            grids = [(factor_x, factor_y, int(number / (factor_x * factor_y)))
                     for factor_x in factors
                     for factor_y in factors
                     if (number / (factor_x * factor_y)).is_integer()]
        return grids
    return []


def get_incomplete_factors(number, factors):
    """
    TODO
    """
    grids = []
    sets = []
    for factor_x in factors:
        for factor_y in factors:
            value = number / (factor_x * factor_y)
            if value.is_integer():
                grid = (factor_x, factor_y, int(value))
                if set(grid) not in sets:
                    sets.append(set(grid))
                    grids.append(grid)
    return grids


def dict_merge(old_dct, merge_dct, add_keys=True, extend_lists=True):
    """ Recursive dict merge. Inspired by :meth:``dict.update()``, instead of
    updating only top-level keys, dict_merge recurses down into dicts nested
    to an arbitrary depth, updating keys. The ``merge_dct`` is merged into
    ``dct``.

    This version will return a copy of the dictionary and leave the original
    arguments untouched.

    The optional argument ``add_keys``, determines whether keys which are
    present in ``merge_dct`` but not ``dct`` should be included in the
    new dict.

    Args:
        old_dct (dict) onto which the merge is executed
        merge_dct (dict): dct merged into dct
        add_keys (bool): whether to add new keys
        extend_lists (bool) : wether to extend lists if keys are updated
                              and value is a list

    Returns:
        dict: updated dict
    """
    dct = deepcopy(old_dct)
    if not add_keys:
        merge_dct = {k: merge_dct[k]
                     for k in set(dct).intersection(set(merge_dct))}

    for key, value in merge_dct.items():
        if isinstance(dct.get(key), dict)\
                and isinstance(value, collections.Mapping):
            dct[key] = dict_merge(dct[key], merge_dct[key],
                                  add_keys=add_keys, extend_lists=extend_lists)
        elif isinstance(dct.get(key), list) and extend_lists:
            dct[key].extend(value)
        else:
            dct[key] = value

    return dct


def is_jsonable(x):
    try:
        json.dumps(x)
        return True
    except:
        return False


def stringify_dict_keys(obj):
    if isinstance(obj, (list, tuple)):
        new_obj = []
        for elt in obj:
            new_obj.append(stringify_dict_keys(elt))

    elif isinstance(obj, dict):
        new_obj = {}
        for key, value in obj.items():
            new_obj[str(key)] = stringify_dict_keys(value)
    else:
        return obj
    return new_obj


def serialize_dict(dict_):
    serialized_dict = {}
    for key, value in dict_.items():
        if hasattr(value, 'to_dict'):
            serialized_value = value.to_dict()
        elif isinstance(value, dict):
            serialized_value = serialize_dict(value)
        elif isinstance(value, (list, tuple)):
            serialized_value = serialize_sequence(value)
        else:
            if not is_jsonable(value):
                msg = 'Attribute {} of value {} is not json serializable'
                raise SerializationError(msg.format(key, value))
            serialized_value = value
        serialized_dict[key] = serialized_value
    return serialized_dict


def serialize_sequence(seq):
    serialized_sequence = []
    for value in seq:
        if hasattr(value, 'to_dict'):
            serialized_sequence.append(value.to_dict())
        elif isinstance(value, dict):
            serialized_sequence.append(serialize_dict(value))
        elif isinstance(value, (list, tuple)):
            serialized_sequence.append(serialize_sequence(value))
        else:
            serialized_sequence.append(value)
    return serialized_sequence


def get_python_class_from_class_name(full_class_name):
    module_name, class_name = full_class_name.rsplit('.', 1)
    module = import_module(module_name)
    class_ = getattr(module, class_name)
    return class_


def dict_to_object(dict_, class_=None, force_generic: bool = False):
    working_dict = dict_.copy()
    if class_ is None and 'object_class' in working_dict:
        class_ = get_python_class_from_class_name(working_dict['object_class'])

    if class_ is not None and issubclass(class_, DessiaObject):
        different_methods = (class_.dict_to_object.__func__
                             is not DessiaObject.dict_to_object.__func__)
        if different_methods and not force_generic:
            obj = class_.dict_to_object(dict_)
            return obj

        if class_._init_variables is None:
            class_argspec = inspect.getfullargspec(class_)
            init_dict = {k: v for k, v in working_dict.items()
                         if k in class_argspec.args}
        else:
            init_dict = {k: v for k, v in working_dict.items()
                         if k in class_._init_variables}
        # TOCHECK Class method to generate init_dict ??
    else:
        init_dict = working_dict

    subobjects = {}
    for key, value in init_dict.items():
        if isinstance(value, dict):
            subobjects[key] = dict_to_object(value)
        elif isinstance(value, (list, tuple)):
            subobjects[key] = sequence_to_objects(value)
        else:
            subobjects[key] = value

    if class_ is not None:
        obj = class_(**subobjects)
    else:
        obj = subobjects
    return obj


def list_hash(list_):
    hash_ = 0
    for element in list_:
        if isinstance(element, list):
            hash_ += list_hash(element)
        elif isinstance(element, dict):
            hash_ += dict_hash(element)
        elif isinstance(element, str):
            hash_ += sum([ord(e) for e in element])
        else:
            hash_ += hash(element)
    return hash_


def dict_hash(dict_):
    hash_ = 0
    for key, value in dict_.items():
        if isinstance(value, list):
            hash_ += list_hash(value)
        elif isinstance(value, dict):
            hash_ += dict_hash(value)
        else:
            hash_ += hash(key) + hash(value)
    return hash_


def sequence_to_objects(sequence):
    # TODO: rename to deserialize sequence? Or is this a duplicate ?
    deserialized_sequence = []
    for element in sequence:
        if isinstance(element, dict):
            deserialized_sequence.append(dict_to_object(element))
        elif isinstance(element, (list, tuple)):
            deserialized_sequence.append(sequence_to_objects(element))
        else:
            deserialized_sequence.append(element)
    return deserialized_sequence


def getdeepattr(obj, attr):
    return reduce(getattr, [obj] + attr.split('.'))


def full_classname(object_, compute_for: str = 'instance'):
    if compute_for == 'instance':
        return object_.__class__.__module__ + '.' + object_.__class__.__name__
    elif compute_for == 'class':
        return object_.__module__ + '.' + object_.__name__
    else:
        msg = 'Cannot compute {} full classname for object {}'
        raise NotImplementedError(msg.format(compute_for, object_))


def serialization_test(obj):
    # TODO: debug infinite recursion? Should we remove thhis ?
    d = obj.to_dict()
    obj2 = obj.dict_to_object(d)
    if obj != obj2:
        msg = 'Object in no more equal to himself '
        msg += 'after serialization/deserialization!'
        raise ModelError(msg)


def deepcopy_value(value, memo):
    # Escaping unhashable types (list) that would be handled after
    try:
        if value in memo:
            return memo[value]
    except TypeError:
        pass

    if isinstance(value, type):  # For class
        return value
    elif hasattr(value, '__deepcopy__'):
        try:
            copied_value = value.__deepcopy__(memo)
        except TypeError:
            copied_value = value.__deepcopy__()
        memo[value] = copied_value
        return copied_value
    else:
        if isinstance(value, list):
            copied_list = []
            for v in value:
                cv = deepcopy_value(v, memo=memo)
                try:
                    memo[v] = cv
                except TypeError:
                    pass
                copied_list.append(cv)
            return copied_list
        elif isinstance(value, dict):
            copied_dict = {}
            for k, v in value.items():
                copied_k = deepcopy_value(k, memo=memo)
                copied_v = deepcopy_value(v, memo=memo)
                try:
                    memo[k] = copied_k
                except TypeError:
                    pass
                try:
                    memo[v] = copied_v
                except TypeError:
                    pass
                copied_dict[copied_k] = copied_v
            return copied_dict
        else:
            new_value = copy.deepcopy(value, memo=memo)
            memo[value] = new_value
            return new_value


def type_fullname(arg):
    if arg.__module__ == 'builtins':
        full_argname = '__builtins__.' + arg.__name__
    else:
        full_argname = serialize_typing(arg)
    return full_argname


def serialize_typing(typing_):
    if is_typing(typing_):
        origin = get_origin(typing_)
        args = get_args(typing_)
        if origin is Union:
            if len(args) == 2 and type(None) in args:
                # This is a false Union => Is a default value set to None
                return serialize_typing(args[0])
            else:
                # Types union
                argnames = ', '.join([type_fullname(a) for a in args])
                return 'Union[{}]'.format(argnames)
        elif origin is list:
            return 'List[' + type_fullname(args[0]) + ']'
        elif origin is tuple:
            argnames = ', '.join([type_fullname(a) for a in args])
            return 'Tuple[{}]'.format(argnames)
        elif origin is dict:
            key_type = type_fullname(args[0])
            value_type = type_fullname(args[1])
            return 'Dict[{}, {}]'.format(key_type, value_type)
        elif origin is InstanceOf:
            return 'InstanceOf[{}]'.format(type_fullname(args[0]))
        else:
            msg = 'Serialization of typing {} is not implemented'
            raise NotImplementedError(msg.format(typing_))
    if isinstance(typing_, type):
        return full_classname(typing_, compute_for='class')
    return str(typing_)
    # raise NotImplementedError('{} of type {}'.format(typing_, type(typing_)))


def type_from_argname(argname):
    splitted_argname = argname.rsplit('.', 1)
    if argname:
        if splitted_argname[0] != '__builtins__':
            return get_python_class_from_class_name(argname)
        # TODO Check for dangerous eval
        return eval(splitted_argname[1])
    return Any


def deserialize_typing(serialized_typing):
    # TODO : handling recursive deserialization
    if isinstance(serialized_typing, str):
        # TODO other builtins should be implemented
        if serialized_typing in ['float', 'builtins.float']:
            return float

        splitted_type = serialized_typing.split('[')
        full_argname = splitted_type[1].split(']')[0]
        if splitted_type[0] == 'List':
            return List[type_from_argname(full_argname)]
        elif splitted_type[0] == 'Tuple':
            if ', ' in full_argname:
                args = full_argname.split(', ')
                if len(args) == 0:
                    return Tuple
                elif len(args) == 1:
                    type_ = type_from_argname(args[0])
                    return Tuple[type_]
                elif len(set(args)) == 1:
                    type_ = type_from_argname(args[0])
                    return Tuple[type_, ...]
                else:
                    msg = ("Heterogenous tuples are forbidden as types for"
                           "workflow non-block variables.")
                    raise TypeError(msg)
            return Tuple[type_from_argname(full_argname)]
        elif splitted_type[0] == 'Dict':
            args = full_argname.split(', ')
            key_type = type_from_argname(args[0])
            value_type = type_from_argname(args[1])
            return Dict[key_type, value_type]
        # elif splitted_type[0] == 'Union':
        #     args = full_argname.split(', ')
    raise NotImplementedError('{}'.format(serialized_typing))


def serialize(deserialized_element):
    if isinstance(deserialized_element, DessiaObject):
        serialized = deserialized_element.to_dict()
    elif isinstance(deserialized_element, dict):
        serialized = serialize_dict(deserialized_element)
    elif is_sequence(deserialized_element):
        serialized = serialize_sequence(deserialized_element)
    else:
        serialized = deserialized_element
    return serialized


def deserialize(serialized_element):
    if isinstance(serialized_element, dict):
        element = dict_to_object(serialized_element)
    elif is_sequence(serialized_element):
        element = sequence_to_objects(serialized_element)
    else:
        element = serialized_element
    return element


def enhanced_deep_attr(obj, sequence):
    """
    Get deep attribute where Objects, Dicts and Lists
    can be found in recursion.

    :param obj: Parent object in which recursively find attribute
                represented by sequence
    :param sequence: List of strings and integers that represents
                     path to deep attribute.
    :return: Value of deep attribute
    """
    if isinstance(sequence, str):
        # Sequence is a string and not a sequence of deep attributes
        if '/' in sequence:
            # Is deep attribute reference
            sequence = deepattr_to_sequence(sequence)
            return enhanced_deep_attr(obj=obj, sequence=sequence)
        # Is direct attribute
        return enhanced_get_attr(obj=obj, attr=sequence)

    # Get direct attrivute
    subobj = enhanced_get_attr(obj=obj, attr=sequence[0])
    if len(sequence) > 1:
        # Recursively get deep attributes
        subobj = enhanced_deep_attr(obj=subobj, sequence=sequence[1:])
    return subobj


def enhanced_get_attr(obj, attr):
    """
    Safely get attribute in obj.
    Obj can be of Object, Dict, or List type

    :param obj: Parent object in which find given attribute
    :param attr: String or integer that represents
                      name or index of attribute
    :return: Value of attribute
    """
    try:
        return getattr(obj, attr)
    except (TypeError, AttributeError):
        try:
            return obj[attr]
        except TypeError:
            classname = obj.__class__.__name__
            msg = "'{}' object has no attribute '{}'.".format(classname, attr)
            track = tb.format_exc()
            raise DeepAttributeError(message=msg, traceback_=track)


def concatenate_attributes(prefix, suffix, type_: str = 'str'):
    wrong_prefix_format = 'Attribute prefix is wrongly formatted.'
    wrong_prefix_format += 'Is of type {}. Should be str or list'
    if type_ == 'str':
        if isinstance(prefix, str):
            return prefix + '/' + str(suffix)
        elif is_sequence(prefix):
            return sequence_to_deepattr(prefix) + '/' + str(suffix)
        else:
            raise TypeError(wrong_prefix_format.format(type(prefix)))
    elif type_ == 'sequence':
        if isinstance(prefix, str):
            return [prefix, suffix]
        elif is_sequence(prefix):
            return prefix + [suffix]
        else:
            raise TypeError(wrong_prefix_format.format(type(prefix)))
    else:
        wrong_concat_type = 'Type {} for concatenation is not supported.'
        wrong_concat_type += 'Should be "str" or "sequence"'
        raise ValueError(wrong_concat_type.format(type_))


def deepattr_to_sequence(deepattr: str):
    sequence = deepattr.split('/')
    healed_sequence = []
    for i, attribute in enumerate(sequence):
        try:
            healed_sequence.append(int(attribute))
        except ValueError:
            healed_sequence.append(attribute)
    return healed_sequence


def sequence_to_deepattr(sequence):
    healed_sequence = [str(attr) if isinstance(attr, int) else attr
                       for attr in sequence]
    return '/'.join(healed_sequence)


def is_bounded(filter_: Filter, value: float):
    bounded = True
    operator = filter_['operator']
    bound = filter_['bound']

    if operator == 'lte' and value > bound:
        bounded = False
    if operator == 'gte' and value < bound:
        bounded = False

    if operator == 'lt' and value >= bound:
        bounded = False
    if operator == 'gt' and value <= bound:
        bounded = False

    if operator == 'eq' and value != bound:
        bounded = False
    return bounded


def is_sequence(obj):
    """
    :param obj: Object to check
    :return: bool. True if object is a sequence but not a string.
                   False otherwise
    """
    return isinstance(obj, collections.abc.Sequence)\
        and not isinstance(obj, str)


def is_builtin(type_):
    return type_ in TYPING_EQUIVALENCES


def type_from_annotation(type_, module):
    """
    Clean up a proposed type if there are stringified
    """
    if isinstance(type_, str):
        # Evaluating types
        if type_ in TYPES_FROM_STRING:
            type_ = TYPES_FROM_STRING[type_]
        else:
            # Evaluating
            type_ = getattr(import_module(module), type_)
    return type_


def is_typing(object_: Any):
    has_module = hasattr(object_, '__module__')
    if has_module:
        in_typings = object_.__module__ in ['typing', 'dessia_common.typings']
    else:
        return False
    has_origin = hasattr(object_, '__origin__')
    has_args = hasattr(object_, '__args__')
    return has_module and has_origin and has_args and in_typings


def jsonschema_from_annotation(annotation, jsonschema_element,
                               order, editable=None, title=None):
    key, typing_ = annotation
    if isinstance(typing_, str):
        raise ValueError

    if title is None:
        title = prettyname(key)
    if editable is None:
        editable = key not in ['return']

    # Compute base entries
    jsonschema_element[key] = {'title': title, 'editable': editable,
                               'order': order,
                               'python_typing': serialize_typing(typing_)}

    if typing_ in TYPING_EQUIVALENCES.keys():
        # Python Built-in type
        jsonschema_element[key]['type'] = TYPING_EQUIVALENCES[typing_]

    elif is_typing(typing_):
        origin = get_origin(typing_)
        args = get_args(typing_)
        if origin is Union:
            if len(args) == 2 and type(None) in args:
                # This is a false Union => Is a default value set to None
                ann = (key, args[0])
                jsonschema_element = jsonschema_from_annotation(
                    annotation=ann, jsonschema_element=jsonschema_element,
                    order=order, editable=editable, title=title
                )
            else:
                # Types union
                classnames = [full_classname(object_=a, compute_for='class')
                              for a in args]

                standalone_args = [a._standalone_in_db for a in args]
                if all(standalone_args):
                    standalone = True
                elif not any(standalone_args):
                    standalone = False
                else:
                    msg = "standalone_in_db values for type '{}'" \
                          " are not consistent"
                    raise ValueError(msg.format(typing_))
                jsonschema_element[key].update({
                    'type': 'object', 'classes': classnames,
                    'standalone_in_db': standalone
                })
        elif origin is list:
            # Homogenous sequences
            jsonschema_element[key].update(jsonschema_sequence_recursion(
                value=typing_, order=order, title=title, editable=editable
            ))
        elif origin is tuple:
            # Heterogenous sequences (tuples)
            items = []
            for type_ in args:
                items.append({'type': TYPING_EQUIVALENCES[type_]})
            jsonschema_element[key].update({'additionalItems': False,
                                            'type': 'array', 'items': items})
        elif origin is dict:
            # Dynamially created dict structure
            key_type, value_type = args
            if key_type != str:
                # !!! Should we support other types ? Numeric ?
                raise NotImplementedError('Non strings keys not supported')
            jsonschema_element[key].update({
                'type': 'object',
                'patternProperties': {
                    '.*': {
                        'type': TYPING_EQUIVALENCES[value_type]
                    }
                }
            })
        elif origin is Subclass:
            warnings.simplefilter('once', DeprecationWarning)
            msg = "\n\nTyping of attribute '{0}' from class {1} "\
                  "uses Subclass which is deprecated."\
                  "\n\nUse 'InstanceOf[{2}]' instead of 'Subclass[{2}]'.\n"
            arg = args[0].__name__
            warnings.warn(msg.format(key, args[0], arg), DeprecationWarning)
            # Several possible classes that are subclass of another one
            class_ = args[0]
            classname = full_classname(object_=class_, compute_for='class')
            jsonschema_element[key].update({
                'type': 'object', 'instance_of': classname,
                'standalone_in_db': class_._standalone_in_db
            })
        elif origin is InstanceOf:
            # Several possible classes that are subclass of another one
            class_ = args[0]
            classname = full_classname(object_=class_, compute_for='class')
            jsonschema_element[key].update({
                'type': 'object', 'instance_of': classname,
                'standalone_in_db': class_._standalone_in_db
            })
        else:
            msg = "Jsonschema computation of typing {} is not implemented"
            raise NotImplementedError(msg.format(typing_))
    elif hasattr(typing_, '__origin__') and typing_.__origin__ is type:
        jsonschema_element[key].update({
            'type': 'object', 'is_class': True,
            'properties': {'name': {'type': 'string'}}
        })
    elif issubclass(typing_, Measure):
        ann = (key, float)
        jsonschema_element = jsonschema_from_annotation(
            annotation=ann, jsonschema_element=jsonschema_element,
            order=order, editable=editable, title=title
        )
        jsonschema_element[key]['units'] = typing_.units
    else:
        classname = full_classname(object_=typing_, compute_for='class')
        if issubclass(typing_, DessiaObject):
            # Dessia custom classes
            jsonschema_element[key].update({
                'type': 'object',
                'standalone_in_db': typing_._standalone_in_db
            })
        else:
            # Statically created dict structure
            jsonschema_element[key].update(static_dict_jsonschema(typing_))
        jsonschema_element[key]['classes'] = [classname]
    return jsonschema_element


def jsonschema_sequence_recursion(value, order: int, title: str = None,
                                  editable: bool = False):
    if title is None:
        title = 'Items'
    jsonschema_element = {'type': 'array', 'order': order,
                          'python_typing': serialize_typing(value)}

    items_type = get_args(value)[0]
    if is_typing(items_type) and get_origin(items_type) is list:
        jss = jsonschema_sequence_recursion(value=items_type, order=0,
                                            title=title, editable=editable)
        jsonschema_element['items'] = jss
    else:
        annotation = ('items', items_type)
        jss = jsonschema_from_annotation(annotation=annotation,
                                         jsonschema_element=jsonschema_element,
                                         order=0, title=title)
        jsonschema_element.update(jss)
    return jsonschema_element


def prettyname(namestr):
    pretty_name = ''
    if namestr:
        strings = namestr.split('_')
        for i, string in enumerate(strings):
            if len(string) > 1:
                pretty_name += string[0].upper() + string[1:]
            else:
                pretty_name += string
            if i < len(strings) - 1:
                pretty_name += ' '
    return pretty_name


def static_dict_jsonschema(typed_dict, title=None):
    warnings.simplefilter('once', DeprecationWarning)
    msg = "\n\nStatic Dict typing is not fully supported.\n" \
          "This will most likely lead to non predictable behavior" \
          " or malfunctionning features. \n" \
          "Define a custom non-standalone class for type '{}'\n\n"
    classname = full_classname(typed_dict, compute_for='class')
    warnings.warn(msg.format(classname), DeprecationWarning)
    jsonschema_element = deepcopy(JSONSCHEMA_HEADER)
    jss_properties = jsonschema_element['properties']

    # Every value is required in a StaticDict
    annotations = get_type_hints(typed_dict)
    jsonschema_element['required'] = list(annotations.keys())

    # TOCHECK : Not actually ordered !
    for i, annotation in enumerate(annotations.items()):
        attribute, typing_ = annotation
        if not is_builtin(typing_):
            msg = "Complex structure as Static dict values is not supported." \
                  "\n Attribute {} got type {} instead of builtin." \
                  "\n Consider creating a custom class for complex structures."
            raise TypeError(msg.format(attribute, typing_))
        jss = jsonschema_from_annotation(annotation=annotation,
                                         jsonschema_element=jss_properties,
                                         order=i, title=title)
        jss_properties.update(jss)
    return jsonschema_element


def set_default_value(jsonschema_element, key, default_value):
    datatype = datatype_from_jsonschema(jsonschema_element[key])
    if default_value is None\
            or datatype in ['builtin', 'heterogeneous_sequence',
                            'static_dict', 'dynamic_dict']:
        jsonschema_element[key]['default_value'] = default_value
    # elif datatype == 'builtin':
    #     jsonschema_element[key]['default_value'] = default_value
    # elif datatype == 'heterogeneous_sequence':
    #     jsonschema_element[key]['default_value'] = default_value
    elif datatype == 'homogeneous_sequence':
        msg = 'Object {} of type {} is not supported as default value'
        type_ = type(default_value)
        raise NotImplementedError(msg.format(default_value, type_))
    elif datatype in ['standalone_object', 'embedded_object',
                      'instance_of', 'union']:
        object_dict = default_value.to_dict()
        jsonschema_element[key]['default_value'] = object_dict
    return jsonschema_element
    # if isinstance(default_value, tuple(TYPING_EQUIVALENCES.keys())) \
    #         or default_value is None:
    #     jsonschema_element[key]['default_value'] = default_value
    # elif is_sequence(default_value):
    #     if datatype == 'heterogeneous_sequence':
    #         jsonschema_element[key]['default_value'] = default_value
    #     else:
    #         msg = 'Object {} of type {} is not supported as default value'
    #         type_ = type(default_value)
    #         raise NotImplementedError(msg.format(default_value, type_))
    # else:
        # if datatype in ['standalone_object', 'embedded_object',
        #                 'subclass', 'union']:
        # object_dict = default_value.to_dict()
        # jsonschema_element[key]['default_value'] = object_dict
        # else:


def inspect_arguments(method, merge=False):
    # Find default value and required arguments of class construction
    args_specs = inspect.getfullargspec(method)
    nargs = len(args_specs.args) - 1

    if args_specs.defaults is not None:
        ndefault_args = len(args_specs.defaults)
    else:
        ndefault_args = 0

    default_arguments = {}
    arguments = []
    for iargument, argument in enumerate(args_specs.args[1:]):
        if argument not in _FORBIDDEN_ARGNAMES:
            if iargument >= nargs - ndefault_args:
                default_value = args_specs.defaults[ndefault_args - nargs
                                                    + iargument]
                if merge:
                    arguments.append((argument, default_value))
                else:
                    default_arguments[argument] = default_value
            else:
                arguments.append(argument)
    return arguments, default_arguments


def deserialize_argument(type_, argument):
    if is_typing(type_):
        origin = get_origin(type_)
        args = get_args(type_)
        if origin is Union:
            # Type union
            classes = list(args)
            instantiated = False
            while instantiated is False:
                # Find the last class in the hierarchy
                hierarchy_lengths = [len(cls.mro()) for cls in classes]
                max_length = max(hierarchy_lengths)
                children_class_index = hierarchy_lengths.index(max_length)
                children_class = classes[children_class_index]
                try:
                    # Try to deserialize
                    # Throws KeyError if we try to put wrong dict into
                    # dict_to_object. This means we try to instantiate
                    # a children class with a parent dict_to_object
                    deserialized_arg = children_class.dict_to_object(argument)

                    # If it succeeds we have the right
                    # class and instantiated object
                    instantiated = True
                except KeyError:
                    # This is not the right class, we should go see the parent
                    classes.remove(children_class)
        elif origin is list:
            # Homogenous sequences (lists)
            sequence_subtype = args[0]
            deserialized_arg = [deserialize_argument(sequence_subtype, arg)
                                for arg in argument]
        elif origin is tuple:
            # Heterogenous sequences (tuples)
            deserialized_arg = tuple([deserialize_argument(t, arg)
                                      for t, arg in zip(args, argument)])
        elif origin is dict:
            # Dynamic dict
            deserialized_arg = argument
        else:
            msg = "Deserialization of typing {} is not implemented"
            raise NotImplementedError(msg.format(type_))
    else:
        if type_ in TYPING_EQUIVALENCES.keys():
            if isinstance(argument, type_):
                deserialized_arg = argument
            else:
                if isinstance(argument, int) and type_ == float:
                    # Explicit conversion in this case
                    deserialized_arg = float(argument)
                else:
                    msg = 'Given built-in type and argument are incompatible: '
                    msg += '{} and {} in {}'.format(type(argument),
                                                    type_, argument)
                    raise TypeError(msg)
        elif issubclass(type_, DessiaObject):
            # Custom classes
            deserialized_arg = type_.dict_to_object(argument)
        else:
            # Static Dict
            deserialized_arg = argument
    return deserialized_arg


# TODO recursive_type and recursive_type functions look weird
def recursive_type(obj):
    if isinstance(obj, tuple(list(TYPING_EQUIVALENCES.keys()) + [dict])):
        type_ = TYPES_STRINGS[type(obj)]
    elif isinstance(obj, DessiaObject):
        type_ = obj.__module__ + '.' + obj.__class__.__name__
    elif isinstance(obj, (list, tuple)):
        type_ = []
        for element in obj:
            type_.append(recursive_type(element))
    elif obj is None:
        type_ = None
    else:
        raise NotImplementedError(obj)
    return type_


def recursive_instantiation(type_, value):
    if type_ in TYPES_STRINGS.values():
        return eval(type_)(value)
    elif isinstance(type_, str):
        class_ = get_python_class_from_class_name(type_)
        if inspect.isclass(class_):
            return class_.dict_to_object(value)
        else:
            raise NotImplementedError
    elif isinstance(type_, (list, tuple)):
        return [recursive_instantiation(t, v) for t, v in zip(type_, value)]
    elif type_ is None:
        return value
    else:
        raise NotImplementedError(type_)


def default_sequence(array_jsonschema):
    if is_sequence(array_jsonschema['items']):
        # Tuple jsonschema
        if 'default_value' in array_jsonschema:
            return array_jsonschema['default_value']
        return [default_dict(v) for v in array_jsonschema['items']]
    return None


def datatype_from_jsonschema(jsonschema):
    if jsonschema['type'] == 'object':
        if 'classes' in jsonschema:
            if len(jsonschema['classes']) > 1:
                return 'union'
            if 'standalone_in_db' in jsonschema:
                if jsonschema['standalone_in_db']:
                    return 'standalone_object'
                return 'embedded_object'
            return 'static_dict'
        if 'instance_of' in jsonschema:
            return 'instance_of'
        if 'patternProperties' in jsonschema:
            return 'dynamic_dict'
        if 'method' in jsonschema and jsonschema['method']:
            return 'embedded_object'
        if 'is_class' in jsonschema and jsonschema['is_class']:
            return 'class'

    elif jsonschema['type'] == 'array':
        if 'additionalItems' in jsonschema\
                and not jsonschema['additionalItems']:
            return 'heterogeneous_sequence'
        return 'homogeneous_sequence'

    elif jsonschema['type'] in ['number', 'string', 'boolean']:
        return 'builtin'
    return None


def chose_default(jsonschema):
    datatype = datatype_from_jsonschema(jsonschema)
    if datatype in ['heterogeneous_sequence', 'homogeneous_sequence']:
        return default_sequence(jsonschema)
    elif datatype == 'static_dict':
        return default_dict(jsonschema)
    elif datatype in ['standalone_object', 'embedded_object',
                      'instance_of', 'union']:
        if 'default_value' in jsonschema:
            return jsonschema['default_value']
        return None
    else:
        return None


def default_dict(jsonschema):
    dict_ = {}
    datatype = datatype_from_jsonschema(jsonschema)
    if datatype in ['standalone_object', 'embedded_object', 'static_dict']:
        for property_, jss in jsonschema['properties'].items():
            if 'default_value' in jss:
                dict_[property_] = jss['default_value']
            else:
                dict_[property_] = chose_default(jss)
    else:
        return None
    return dict_


class ParsedAttribute(TypedDict):
    desc: str
    type_: str
    annotation: str


class ParsedDocstring(TypedDict):
    description: str
    attributes: Dict[str, ParsedAttribute]


def parse_docstring(cls: Type) -> ParsedDocstring:
    """
    Parse docstring of given class. Refer to docs to see how docstrings
    should be built.
    """
    annotations = get_type_hints(cls.__init__)
    docstring = cls.__doc__
    if docstring:
        splitted_docstring = docstring.split(':param ')
        parsed_docstring = {"description": splitted_docstring[0].strip()}
        params = splitted_docstring[1:]
        args = {}
        for param in params:
            splitted_param = param.split(':type ')
            arg = splitted_param[0]
            typestr = splitted_param[1]
            argname, argdesc = arg.split(":", maxsplit=1)
            argtype = typestr.split(argname+":")[-1]
            annotation = annotations[argname]
            args[argname] = {'desc': argdesc.strip(), 'type_': argtype.strip(),
                             'annotation': str(annotation)}
            # TODO Should be serialize typing ?
        parsed_docstring.update({'attributes': args})
        return parsed_docstring
    return {'description': "", 'attributes': {}}
