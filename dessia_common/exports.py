#!/usr/bin/env python3
# -*- coding: utf-8 -*-
""" Exports for dessia_common. """
from typing import List, Dict, Any, Sequence, Optional


from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook
import openpyxl.utils
from dessia_common.utils.types import is_sequence
from dessia_common.breakdown import breakdown


def is_hashable(value):
    """ Determine whether `value` can be hashed. """
    try:
        hash(value)
    except TypeError:
        return False
    return True


def is_number(value):
    """ Determine if the value is a int or a float. """
    return isinstance(value, (int, float))


def is_builtins_list(list_):
    """ Determine if a list is only composed of builtins. """
    for element in list_:
        if not (is_number(element) or isinstance(element, str)):
            return False
    return True


class ExportFormat:
    """
    Define which method of an object should be called for each Export.

    :param str selector: A custom and unique identifier for current ExportFormat.
    :param str extension: The extension of the file to create (without '.')
    :param str method_name: The method to call to generate the export. Should be the one that handles the stream.
    :param bool text: Whether the resulting file is text-like or binary-like.
    :param str export_name: Enables user to set their own custom export name.
    :param Dict args: Arguments to pass to the called function.
    """

    def __init__(self, selector: Optional[str], extension: str, method_name: str, text: bool,
                 export_name: str = "", args: Dict[str, Any] = None):
        self.selector = selector
        self.extension = extension
        self.method_name = method_name
        self.text = text
        self.export_name = export_name
        if args is None:
            args = {}
        self.args = args

    def to_dict(self):
        """ Serialization. """
        return {"selector": self.selector, "extension": self.extension, "method_name": self.method_name,
                "text": self.text, "export_name": self.export_name, "args": self.args}


class XLSXWriter:
    """ Base class to write a DessiaObject in an excel file. """

    max_column_width = 40
    color_dessIA1 = "263238"
    color_dessIA2 = "537CB0"
    grey1 = "f1f1f1"
    white_font = Font(color="FFFFFF")

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    def __init__(self, object_):
        self.pattern_color1 = PatternFill(
            fill_type="solid",
            start_color=self.color_dessIA1,
            end_color=self.color_dessIA1)

        self.pattern_color2 = PatternFill(
            fill_type="solid",
            start_color=self.color_dessIA2,
            end_color=self.color_dessIA2)

        self.workbook = Workbook()
        self.main_sheet = self.workbook.active
        self.object = object_

        self.paths = breakdown(object_)

        self.classes_to_sheets = {}
        self.object_to_sheet_row = {}
        for class_name, obj_paths in self.paths.items():
            sheet = self.workbook.create_sheet(class_name)
            self.classes_to_sheets[class_name] = sheet

            for i, (obj, path) in enumerate(obj_paths.items()):
                self.object_to_sheet_row[obj] = (sheet, i + 2, path)

        self.write()

    def write_class_header_to_row(self, obj_of_class, sheet, row_number):
        """ Write to a sheet the class header: finds columns names from a class. """
        cell = sheet.cell(row=row_number, column=1, value='Path')
        cell.fill = self.pattern_color2
        cell.border = self.thin_border
        cell.font = self.white_font

        cell = sheet.cell(row=row_number, column=2, value='name')
        cell.fill = self.pattern_color2
        cell.border = self.thin_border
        cell.font = self.white_font
        i = 3

        for (k, _) in sorted(obj_of_class.__dict__.items()):
            if (not k.startswith('_')) and k != 'name':
                cell = sheet.cell(row=row_number, column=i, value=str(k))

                cell.border = self.thin_border
                cell.fill = self.pattern_color2
                cell.font = self.white_font
                i += 1

    def compute_cell_link(self, value):
        """ Compute the hyperlink for given value."""
        cell_link = None
        if isinstance(value, dict):
            cell_link = f'#{list(value.values())[0].__class__.__name__}!A2'
        elif isinstance(value, (list, set)) and not is_builtins_list(value):
            cell_link = f'#{value[0].__class__.__name__}!A2'
        elif is_hashable(value) and value in self.object_to_sheet_row:
            ref_sheet, ref_row_number, _ = self.object_to_sheet_row[value]
            cell_link = f'#{ref_sheet.title}!A{ref_row_number+2}'

        return cell_link

    def write_value_to_cell(self, value, sheet, row_number, column_number):
        """ Write a given value to a cell. Insert it as a link if it is an object. """
        if isinstance(value, dict):
            cell_content = f'Dict of {len(value)} items'
        elif isinstance(value, list):
            if is_builtins_list(value):
                cell_content = str(value)
            else:
                cell_content = f'List of {len(value)} items'

        elif isinstance(value, set):
            cell_content = f'Set of {len(value)} items'

        elif isinstance(value, float):
            cell_content = round(value, 6)
        elif is_hashable(value) and value in self.object_to_sheet_row:
            _, _, ref_path = self.object_to_sheet_row[value]
            cell_content = ref_path

        else:
            cell_content = str(value)

        cell = sheet.cell(row=row_number, column=column_number, value=cell_content)

        cell_link = self.compute_cell_link(value=value)
        if cell_link:
            cell.hyperlink = cell_link

        cell.border = self.thin_border

    def write_object_to_row(self, obj, sheet, row_number, path=''):
        """ Write on object to a row. Loops on its attributes to write its value in each cell. """
        cell = sheet.cell(row=row_number, column=1, value=path)
        cell.border = self.thin_border
        if hasattr(obj, 'name'):
            cell = sheet.cell(row=row_number, column=2, value=obj.name)
        else:
            cell = sheet.cell(row=row_number, column=2, value='No name in model')

        cell.border = self.thin_border
        i = 3
        for (k, value) in sorted(obj.__dict__.items()):
            if (not k.startswith('_')) and k != 'name':
                self.write_value_to_cell(value, sheet, row_number, i)

                i += 1

                # column_width = min((len(k) + 1.5), max_column_width)
                # column_name = openpyxl.utils.cell.get_column_letter(i)
                # sheet.column_dimensions[column_name].width = column_width

    def write_object_id(self, sheet):
        """ Write object id to a given sheet. """
        sheet.title = f'Object {self.object.__class__.__name__}'

        sheet['A1'] = 'Module'
        sheet['B1'] = 'Class'
        sheet['C1'] = 'name'

        sheet['A1'].border = self.thin_border
        sheet['B1'].border = self.thin_border
        sheet['C1'].border = self.thin_border
        sheet['A1'].fill = self.pattern_color1
        sheet['B1'].fill = self.pattern_color1
        sheet['C1'].fill = self.pattern_color1
        sheet['A1'].font = self.white_font
        sheet['B1'].font = self.white_font
        sheet['C1'].font = self.white_font

        sheet['A2'] = self.object.__module__
        sheet['B2'] = self.object.__class__.__name__
        sheet['C2'] = self.object.name

        sheet['A2'].border = self.thin_border
        sheet['B2'].border = self.thin_border
        sheet['C2'].border = self.thin_border
        sheet['A2'].fill = self.pattern_color1
        sheet['B2'].fill = self.pattern_color1
        sheet['C2'].fill = self.pattern_color1
        sheet['A2'].font = self.white_font
        sheet['B2'].font = self.white_font
        sheet['C2'].font = self.white_font

        sheet['A3'] = 'Attribute'
        sheet['A4'] = 'Value'
        sheet['A3'].border = self.thin_border
        sheet['A4'].border = self.thin_border

    def write(self):
        """ Generate the whole file. """
        # name_column_width = 0
        self.write_object_id(self.main_sheet)
        self.write_class_header_to_row(self.object, self.main_sheet, 3)
        self.write_object_to_row(self.object, self.main_sheet, 4)
        self.autosize_sheet_columns(self.main_sheet, 5, 30)

        for class_name, obj_paths in self.paths.items():
            sheet = self.classes_to_sheets[class_name]

            sheet['A1'] = 'Module'
            sheet['B1'] = 'Class'
            obj_info = list(obj_paths.keys())[0]
            sheet['A2'] = obj_info.__module__
            sheet['B2'] = obj_info.__class__.__name__

            for obj, path in obj_paths.items():
                _, row_number, path = self.object_to_sheet_row[obj]
                self.write_object_to_row(obj, sheet, row_number+2, path)
            self.write_class_header_to_row(obj, sheet, 3)
            sheet.auto_filter.ref = f"A1:{openpyxl.utils.cell.get_column_letter(sheet.max_column)}{len(obj_paths) + 1}"
            self.autosize_sheet_columns(sheet, 5, 30)

    def save_to_file(self, filepath: str):
        """ Save to a filepath (open) and write. """
        if not filepath.endswith('.xlsx'):
            filepath += '.xlsx'
            print(f"Changing name to {filepath}")

        with open(filepath, 'rb') as file:
            self.save_to_stream(file)

    def save_to_stream(self, stream):
        """ Saves the file to a binary stream. """
        self.workbook.save(stream)

    @staticmethod
    def autosize_sheet_columns(sheet, min_width=5, max_width=30):
        """ Auto-size the sheet columns by analyzing the content. Min and max width must be specified. """
        for col in sheet.columns:
            width = min_width
            column = col[1].column_letter  # Get the column name
            # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based)
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > width:
                        width = len(str(cell.value))
                except AttributeError:
                    pass
            if width > 0:
                adjusted_width = min((width + 0.5), max_width)
                sheet.column_dimensions[column].width = adjusted_width


class MarkdownWriter:
    """ Base class to write markdowns. """

    def __init__(self, print_limit: int = 25, table_limit: Optional[int] = 12):
        self.print_limit = print_limit
        self.table_limit = table_limit

    @staticmethod
    def object_titles():
        """ Return a list of strings representing the titles for the object matrix. """
        return ['Attribute', 'Type', 'Value']

    @staticmethod
    def _sequence_to_str(value: Sequence):
        if len(value) == 0:
            return f"empty {type(value).__name__}"

        printed_string = f"{len(value)} "

        all_class_names = list(set(subvalue.__class__.__name__ for subvalue in value))
        if len(all_class_names) == 1:
            printed_string += ''.join([f"{all_class_names[0]}",
                                       f"{'s' if len(value) > 1 else ''}"])
        else:
            str_all_class = str(all_class_names).translate(str(all_class_names).maketrans('', '', "{}'"))
            printed_string += f"{str_all_class}"

        return printed_string

    def _dict_to_str(self, value: Dict):
        return self._sequence_to_str(list(value.values()))

    @staticmethod
    def _object_to_str(value) -> str:
        if hasattr(value, 'name') and value.name:
            return value.name
        return 'unnamed'

    def _value_to_str(self, value: Any) -> str:
        if isinstance(value, (float, int, bool, complex)):
            return str(round(value, 6))
        if isinstance(value, str):
            return value if value != '' else 'no value'
        if is_sequence(value):
            return self._sequence_to_str(value)
        if isinstance(value, Dict):
            return self._dict_to_str(value)
        if value is None:
            return ' - '
        return self._object_to_str(value)

    def _string_in_table(self, string: str = ''):
        return string[:self.print_limit] + ('...' if len(string) > self.print_limit else '')

    def object_matrix(self, object_):
        """ Return a matrix representing the object passed as argument. """
        matrix = []
        for attr, value in object_.__dict__.items():
            matrix.append([attr,
                           value.__class__.__name__,
                           self._value_to_str(value)])
        return matrix

    @staticmethod
    def _head_table(col_names: List[str]) -> str:
        cap_names = map(lambda x: x.capitalize(), col_names)
        return ("| " + " | ".join(cap_names) + " |\n" +
                "| ------ " * len(col_names) + "|\n")

    def _table_line(self, row: List[Any]) -> str:
        line = "|"
        for value in row:
            line += f" {self._string_in_table(self._value_to_str(value))} |"
        return line + "\n"

    def _table_rows_from_content(self, content: List[List[Any]]) -> str:
        string = ''
        for row in content:
            string += self._table_line(row)
        return string

    def _content_table(self, content: List[List[Any]]) -> str:
        if self.table_limit is None:
            return self._table_rows_from_content(content)

        table = ''
        half_table = int(self.table_limit / 2)

        table += self._table_rows_from_content(content[:half_table])

        if self.table_limit > 1:
            if len(content) > self.table_limit:
                table += f"| + {len(content) - self.table_limit} unprinted elements | |\n"

            if len(content) > half_table:
                table += self._table_rows_from_content(content[-half_table:])

        return table

    @staticmethod
    def print_name(object_) -> str:
        """Print name of object_."""
        return object_.name if object_.name != '' else 'with no name'

    @staticmethod
    def print_class(object_) -> str:
        """Print name of class name of object_."""
        return object_.__class__.__name__

    def matrix_table(self, matrix: List[List[float]], col_names: List[str]) -> str:
        """Print col_names of matrix as a table."""
        return ''.join([self._head_table(col_names),
                        self._content_table(matrix)])

    def object_table(self, object_) -> str:
        """Print object_'s attributes in table."""
        return self.matrix_table(self.object_matrix(object_),
                                 self.object_titles())

    def element_details(self, elements: List[Any]) -> str:
        """Print sequence of elements."""
        return self._sequence_to_str(elements)

    @staticmethod
    def write_to_file(filename: str, content: str) -> None:
        """ Writes the given content to the specified file. """
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(content)

    @staticmethod
    def table_of_contents(headings: List[str]) -> str:
        """ Generates a table of contents based on the given list of headings. """
        table_of_contents = '## Table of Contents\n\n'
        for heading in headings:
            table_of_contents += f'- [{heading}](#{heading.lower().replace(" ", "-")})\n'
        return table_of_contents

    @staticmethod
    def header(title: str, level: int = 1) -> str:
        """ Generates a markdown header with the specified title and level. """
        header_level = "#" * level
        return f"{header_level} {title}\n\n"
